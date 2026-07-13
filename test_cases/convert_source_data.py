"""
test_cases/convert_source_data.py — DONNEE~1.XLS -> test_cases/data/<ticker>.csv
=================================================================================
DONNEE~1.XLS (racine du repo) a une extension trompeuse : son contenu réel est un
zip .xlsx (vérifié via les octets magiques PK\x03\x04), mais openpyxl refuse de
l'ouvrir sur la seule extension ".xls" (ancien format binaire OLE2, non supporté).
On copie donc le fichier vers un temporaire ".xlsx" avant lecture — aucune donnée
n'est modifiée, seule l'extension du chemin lu diffère.

Chaque feuille (une par actif, nom = "label" de calibration/regime/assets.py, ex.
"Bitcoin", "S&P 500 (SPY)") contient les colonnes :
    Date, Close, Sigma_t_pct, Vol_of_Vol, Volume_norm, Changepoint_prob, Regime
déjà calculées par le moteur calibration/regime (HMM + GARCH + ADX). On les
réécrit telles quelles en CSV sous test_cases/data/, pour que le reste du
pipeline de test cases n'ait plus jamais besoin d'Excel ni du fichier source à
l'exécution.

Exécution (depuis DeepEdgeBenchmark/) :
    python -m test_cases.convert_source_data
"""

import shutil
import tempfile
from pathlib import Path

import openpyxl
import pandas as pd

from test_cases.registry_assets import ASSETS

REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLS = REPO_ROOT / "DONNEE~1.XLS"

COLUMNS = ["Date", "Close", "Sigma_t_pct", "Vol_of_Vol", "Volume_norm", "Changepoint_prob", "Regime"]


def _read_sheet(wb, sheet_name: str) -> pd.DataFrame:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    df = pd.DataFrame(rows, columns=COLUMNS)
    df["Date"] = pd.to_datetime(df["Date"])
    return df.sort_values("Date").reset_index(drop=True)


def convert(source_xls: Path = SOURCE_XLS, out_dir: Path = None) -> list:
    """Convertit chaque feuille de source_xls en CSV sous out_dir (défaut :
    test_cases/data/). Retourne la liste des chemins CSV écrits."""
    out_dir = out_dir or (Path(__file__).parent / "data")
    out_dir.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory() as tmp:
        tmp_xlsx = Path(tmp) / "source.xlsx"
        shutil.copy2(source_xls, tmp_xlsx)
        wb = openpyxl.load_workbook(str(tmp_xlsx), data_only=True, read_only=True)

        written = []
        for asset in ASSETS:
            sheet_name = asset["sheet_name"]
            if sheet_name not in wb.sheetnames:
                print(f"[convert_source_data] feuille absente pour {asset['ticker']!r} : {sheet_name!r} — ignoré")
                continue
            df = _read_sheet(wb, sheet_name)
            csv_path = out_dir / asset["csv_path"].name
            df.to_csv(csv_path, index=False)
            written.append(csv_path)
            print(f"[convert_source_data] {asset['ticker']:<8} {sheet_name:<32} -> {csv_path} ({len(df)} lignes)")

        wb.close()  # libère le handle sur tmp_xlsx avant le nettoyage du TemporaryDirectory (Windows)

    return written


if __name__ == "__main__":
    convert()
